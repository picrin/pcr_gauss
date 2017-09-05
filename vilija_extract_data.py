from __future__ import print_function
import openpyxl
import sys

print(sys.version)

book = openpyxl.load_workbook('vilija_data.xlsm')

sheetnames = book.sheetnames[1:]


print(sheetnames)

#count = 0
for sheetname in sheetnames:
    sheet = book[sheetname]
    mr = sheet.max_row

    with open("data/" + sheetname + ".tsv", "w") as s:

        for i in range(1, mr):
            row = sheet[i + 1]
            print(row[0].value, file=s, end="\t")
            print(row[3].value, file=s)